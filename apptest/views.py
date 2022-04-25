import json
import os
import random
from datetime import datetime

import requests
from django import forms
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import RegexValidator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt

from apptest.models import UserInfo, Department
from apptest import models
from apptest.utils.pagination import Pagination
from apptest.utils.bootstrap import BootstrapModelForm
from apptest.utils.encrypt import md5

def newwatch(request):
    import requests
    res = requests.get('https://news.cctv.com/2019/07/gaiban/cmsdatainterface/page/news_1.jsonp?cb=news')
    newdata=res.json()
    print(newdata)
    return  render(request,'news.html',{'newstitle':newdata})

def adminsth(request):
    return HttpResponse("adminsth页面内容")

def login(request):
    if request.method == 'GET':
        return render(request,'login.html')
    print(request.POST)
    username = request.POST.get('username')
    pwd =request.POST.get('pwd')
    if username == 'root' and pwd =='123':
        return redirect('http://www.chinaunicom.com.cn/index.html')
    else:
        return render(request,'login.html',{'errormsg':"用户名或密码错误"})

def info_list(request):
    # users=UserInfo.objects.all()
    users = UserInfo.objects.all()
    return render(request, 'info_list.html', {"userItems": users})

def info_add(request):
    if request.method=='GET':
        return render(request,'info_add.html')

    setname = request.POST.get('username')
    setpwd = request.POST.get('password')
    setage = request.POST.get('age')
    UserInfo.objects.create(name=setname, password=setpwd, age=setage)
    return redirect('/info_list/')

def info_delete(request):
    nid=request.GET.get('nid')
    UserInfo.objects.filter(id=nid).delete()
    return redirect('/info_list/')

def depart_info(request):
    queryset=models.Department.objects.all()
    return render(request,"depart_info.html",{"queryset":queryset})

def depart_infoadd(request):
    if request.method=='GET':
      return render(request,"depart_infoadd.html")
    title=request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect('/depart_info/')

def depart_infodel(request):
    pid=request.GET.get('pid')
    models.Department.objects.filter(id=pid).delete()
    return redirect('/depart_info/')

def depart_infoedit(request,nid):
    if  request.method == "GET":
        partedit=models.Department.objects.filter(id=nid).first()
        return render(request,"depart_infoedit.html",{"part_title":partedit.title})
    updatetitle = request.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=updatetitle)
    return redirect('/depart_info/')

def users_info(request):
    """用户列表展示"""
    queryset=models.UserInfo.objects.all()
    page_object=Pagination(request,queryset)
    context={
        "queryset":page_object.page_queryset,
        "page_string":page_object.html()
    }
    return render(request,"users_info.html",context)

#以ModelForm实现表单
class UsersModelForm(BootstrapModelForm):
    class Meta:
        model = models.UserInfo
        fields={"name","password","age","account","create_time","gender","depart"}

def users_infoadd(request):
    if request.method=='GET':
        form = UsersModelForm()
        return render(request,"users_infoadd.html",{"form":form})
    #post提交数据校验
    form = UsersModelForm(data=request.POST)
    if form.is_valid():
        #如果数据合法加入数据库
        form.save()
        return redirect('/users_info/')
    return render(request, "users_infoadd.html", {"form": form})

def users_infoedit(request,nid):
    row_object=models.UserInfo.objects.filter(id=nid).first()
    if request.method=='GET':
        form = UsersModelForm(instance=row_object)
        return render(request,"users_infoedit.html",{"form":form})

    form_edit = UsersModelForm(data=request.POST,instance=row_object)
    if form_edit.is_valid:
        #默认保存用户编辑输入的内容，若需要保存默认值，则可以使用
        #form.instance,字段名=值
        form_edit.save()
        return redirect('/users_info/')
    return render(request, "users_infoedit.html", {"form": form_edit})

def users_infodel(request):
    pid=request.GET.get('pid')
    models.UserInfo.objects.filter(id=pid).delete()
    return redirect('/users_info/')

def phones_info(request):
    search_data_dict={}
    search_data=request.GET.get('q',"")
    if search_data:
        search_data_dict["mobile__contains"] =  search_data

    # 靓号列表展示
    queryset = models.Phone.objects.filter(**search_data_dict).order_by("id")  # 按某字段+升-降排序
    pagination=Pagination(request,queryset)
    page_queryset1=pagination.page_queryset
    page_html_object=pagination.html()
    return render(request,'phones_info.html',{
        "queryset":page_queryset1,
        "search_data":search_data,
        "page_string":page_html_object
    })

class PhoneAddModelForm(BootstrapModelForm):
    # 字段验证方式1：RegexValidator正则
    mobile = forms.CharField(label='靓号',validators=[RegexValidator(r'^1[3-9]\d{9}$','手机格式错误')])
    class Meta:
        model=models.Phone
        fields={"mobile","price","level","status"}

    #加入不重复的手机号
    def clean_mobile(self):
        #获取编辑的那一行记录的ID
        txt_mobile=self.cleaned_data['mobile']
        is_exists=models.Phone.objects.filter(mobile=txt_mobile).exists()
        models.Phone.objects.filter(mobile=txt_mobile).exclude(id=2)
        if is_exists:
            raise ValidationError('手机号已存在')
        return txt_mobile
    #字段验证方式2：对用户输入的值进行验证
    # def clean_mobile(self):
    #     txt_mobile=self.cleaned_data['mobile']
    #     if len(txt_mobile) != 11:
    #         raise ValidationError("格式错误")
    #     #验证通过，用户输入的值返回
    #     return txt_mobile

def phones_infoadd(request):
   if request.method=='GET':
        phone_form=PhoneAddModelForm()
        return render(request,'phones_infoadd.html',{"form":phone_form})

   phone_form=PhoneAddModelForm(data=request.POST)
   if phone_form.is_valid():
       phone_form.save()
       return redirect('/phones_info/')
   return render(request, 'phones_infoadd.html', {"form": phone_form})

class PhoneEditModelForm(forms.ModelForm):
    # 编辑操作时需要对某个字段让其不可编辑
    # mobile = forms.CharField(disabled=True)
    class Meta:
        model=models.Phone
        fields={"mobile","price","level","status"}
    #编辑不重复的手机号
    def clean_mobile(self):
        #获取编辑的那一行记录的ID:self.instance.pk
        txt_mobile=self.cleaned_data['mobile']
        is_exists=models.Phone.objects.exclude(id=self.instance.pk).filter(mobile=txt_mobile).exists()
        if is_exists:
            raise ValidationError('手机号已存在')
        return txt_mobile

def phones_infoedit(request,nid):
    row_object=models.Phone.objects.filter(id=nid).first()
    if request.method=='GET':
        phone_form=PhoneEditModelForm(instance=row_object)
        return render(request,'phones_infoedit.html',{"form":phone_form})
    phone_form = PhoneEditModelForm(instance=row_object,data=request.POST)
    if phone_form.is_valid():
        phone_form.save()
        return redirect('/phones_info/')
    return render(request, 'phones_infoedit.html', {"form": phone_form})

def phones_infodel(request):
    del_pid=request.GET.get("pid")
    models.Phone.objects.filter(id=del_pid).delete()
    return redirect('/phones_info/')

def admins_info(request):
    search_data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        search_data_dict["adminname__contains"] = search_data

    # 管理员列表展示
    queryset = models.Admins.objects.filter(**search_data_dict).order_by("id")  # 按某字段+升-降排序
    pagination = Pagination(request, queryset,page_size=10)
    page_queryset1 = pagination.page_queryset
    page_html_object = pagination.html()
    context={
        "queryset": page_queryset1,
        "search_data": search_data,
        "page_string": page_html_object,
        "title":"新增管理员"
    }
    return render(request, 'admins_info.html',context)

class AdminAddModelForm(BootstrapModelForm):
    comfirm_pwd=forms.CharField(
        label="确认密码",
        max_length=32,
        widget=forms.PasswordInput(render_value=True)
    )
    class Meta:
        model =models.Admins
        fields=['adminname','adminpwd','comfirm_pwd']
        widget={
           "adminpwd":forms.PasswordInput(render_value=True)
        }

    def clean_adminpwd(self):
        txt_adminpwd=self.cleaned_data.get("adminpwd")
        return md5(txt_adminpwd)

    def clean_confire_pwd(self):
        txt_adminpwd=self.cleaned_data["adminpwd"]
        txt_comfire_pwd=md5(self.cleaned_data["comfire_pwd"])
        if txt_comfire_pwd != txt_adminpwd:
            raise ValidationError('密码不一致')
        return txt_comfire_pwd

def admins_infoadd(request):
    title="新建管理员"
    if request.method=='GET':
        form =AdminAddModelForm()
        return render(request,'admins_infoadd.html',{ "title":title,"form": form})

    add_form=AdminAddModelForm(data=request.POST)
    if add_form.is_valid():
        add_form.save()
        return redirect("/admins_info/")
    return render(request, 'admins_infoadd.html', {"title": title, "form": add_form})

class AdminEditModelForm(BootstrapModelForm):
    class Meta:
        model=models.Admins
        fields=['adminname']

def admins_infoedit(request,nid):
    row_object=models.Admins.objects.filter(id=nid).first()
    if not row_object :
        #  return render(request,"error.html",{"msg":数据不存在})
        return redirect("/admins_info/")

    edit_title="编辑管理员"
    if request.method=='GET':
        form = AdminEditModelForm(instance=row_object)
        return render(request, 'admins_infoedit.html',{"form": form,"title":edit_title})

    edit_form=AdminEditModelForm(data=request.POST,instance=row_object)
    if edit_form.is_valid():
        edit_form.save()
        return redirect("/admins_info/")
    return render(request, 'admins_infoedit.html',{"form": edit_form,"title":edit_title})

def admins_infodel(request):
    del_pid=request.GET.get("pid")
    models.Admins.objects.filter(id=del_pid).delete()
    return redirect('/admins_info/')

class AdminSetModelForm(BootstrapModelForm):
    #bug1,确认密码功能无效
    comfirm_pwd=forms.CharField(
        label="确认密码",
        max_length=32,
        widget=forms.PasswordInput(render_value=True)
    )
    class Meta:
        model =models.Admins
        fields=["adminname","adminpwd","comfirm_pwd"]
        widget={
            "adminname":forms.CharField(disabled=True),
            "adminpwd":forms.PasswordInput(render_value=True)
        }

    def clean_adminpwd(self):
        txt_adminpwd=self.cleaned_data.get("adminpwd")
        return md5(txt_adminpwd)

    def clean_confire_pwd(self):
        txt_adminpwd=self.cleaned_data["adminpwd"]
        txt_comfire_pwd=md5(self.cleaned_data["comfire_pwd"])
        if txt_comfire_pwd != txt_adminpwd:
            raise ValidationError("密码不一致")
        return txt_comfire_pwd


def admins_inforeset(request,nid):
    row_obj=models.Admins.objects.filter(id=nid).first()
    title = "重置密码"
    if request.method == 'GET':
        form=AdminSetModelForm(instance=row_obj)
        return render(request,'admins_inforeset.html',{"title":title, "form":form})

    reset_form = AdminSetModelForm(data=request.POST,instance=row_obj)
    if reset_form.is_valid():
        reset_form.save()
        return redirect("/admins_info/")
    return render(request, 'admins_inforeset.html', {"title":title,"form":reset_form})

class LoginModelForm(BootstrapModelForm):
    adminname=forms.CharField(
        widget=forms.TextInput,
        required=True
    )
    adminpwd=forms.CharField(
        widget=forms.PasswordInput(render_value=True),
        required=True
    )
    # #bug1:报错提示无
    class Meta:
        model=models.Admins
        fields=['adminname','adminpwd']

    def clean_adminpwd(self):
        txt_adminpwd=self.cleaned_data.get("adminpwd")
        return md5(txt_adminpwd)

    def clean_adminname(self):
        txt_adminname=self.cleaned_data.get("adminname")
        return txt_adminname

def to_login(request):
    if request.method=='GET':
        form = LoginModelForm()
        return render(request,"to_login.html",{"form":form})

    form=LoginModelForm(data=request.POST)

    if form.is_valid():
        row_obj = models.Admins.objects.filter(**form.cleaned_data).first()
        print("登录用户查询结果",row_obj.id)
        if not row_obj:
            form.add_error('adminpwd',"密码错误")
            return render(request, "to_login.html", {"form": form})
        request.session["uid"]=row_obj.id
        request.session["adminname"]=row_obj.adminname
        request.session.set_expiry(7*60*60*24)
        return redirect('/admins_info/')
    return render(request,"to_login.html",{"form":form})


def to_logout(request):
    request.session.clear()
    return redirect('/to_login/')


def test_ajax(request):
    print(request.GET)
    return HttpResponse('Ajax get请求发送成功')


class TaskAddModelForm(forms.ModelForm):
    class Meta:
        model=models.Task
        fields="__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        print(self.fields.items)
        for name, field in self.fields.items():
            if field.widget.attrs:
                field.widget.attrs["class"]= "form-control"
                field.widget.attrs["placeholder"]= field.label
            else:
                field.widget.attrs={"class":"form-control","placeholder": field.label}

@csrf_exempt
def task_info(request):
    form=TaskAddModelForm()
    return render(request,'task_info.html',{"form":form})

@csrf_exempt
def task_infoadd(request):
    #1，用户发送过来的数据进行验证
    form=TaskAddModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        data_dict = {"status":True}
        return HttpResponse(json.dumps(data_dict))
    data_dict = {"status": False,"error":form.errors}
    return HttpResponse(json.dumps(data_dict,ensure_ascii=False))

class OrderModelForm(forms.ModelForm):
    class Meta:
        model=models.Order
        exclude=["oid","admin"]
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        print(self.fields.items)
        for name, field in self.fields.items():
            if field.widget.attrs:
                field.widget.attrs["class"]= "form-control"
                field.widget.attrs["placeholder"]= field.label
            else:
                field.widget.attrs={"class":"form-control","placeholder": field.label}

def order_info(request):
    queryset=models.Order.objects.all().order_by("-id")
    page_object=Pagination(request,queryset)
    form = OrderModelForm()
    context={
        "form": form,
        "title":"订单列表",
        "queryset":page_object.page_queryset,
        "page_string":page_object.html()
    }

    return render(request,'order_info.html',context)

@csrf_exempt
def order_infoadd(request):
    """新建订单 提交Ajax请求"""
    form=OrderModelForm(data=request.POST)
    if form.is_valid():
        #额外增加一些不是用户输入的值，随机订单号生成
        form.instance.oid=datetime.now().strftime("%Y%m%d%H%M%S")+ str(random.randint(1000,9999))
        #指定管理员为当前登录用户
        form.instance.admin=request.session.uid
        form.save()
        return JsonResponse({"status":True})

    return JsonResponse({"status":False,"error":form.errors})

def order_infodelete(request):
    #删除订单
    deleteid=request.GET.get("pkid")
    exists = models.Order.objects.filter(id=deleteid).exists()
    if not exists:
        return JsonResponse({"status":False,"error":"数据不存在"})
    models.Order.objects.filter(id=deleteid).delete()
    return JsonResponse({"status": True})

def order_infodetail(request):
    #编辑订单
    detailid=request.GET.get("pkid")
    #方式1： 获取编辑订单的数据
    # row_object = models.Order.objects.filter(id=detailid).first()
    # if not row_object:
    #     return  JsonResponse({"status": False, "error": "数据不存在"})
    # result={
    #     "status":True,
    #     "data":{
    #         "title":row_object.title,
    #         "price":row_object.price,
    #         "status":row_object.status
    #     }
    # }
    # return JsonResponse(result)
   #方式2： 获取编辑订单的数据
    row_dict = models.Order.objects.filter(id=detailid).values("title", "price", "status").first()
    if not row_dict:
      return  JsonResponse({"status": False, "error": "数据不存在"})
    result = {
        "status":True,
        "data":row_dict}
    return JsonResponse(result)

@csrf_exempt
def order_infoedit(request):
    #编辑订单
    pkid=request.GET.get("pkid")
    row_object = models.Order.objects.filter(id=pkid).exists()
    if not row_object:
        return JsonResponse({"status": False, "tips": "数据不存在"})
    form = OrderModelForm(data=request.POST,instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status":True})

    return JsonResponse({"status":False,"error":form.errors})

def echarts_info(request):
    """数据统计页面"""
    return render(request,'echarts_info.html')

def echarts_infobar(request):
    """构造柱状图数据"""
    return JsonResponse()

def upload_info(request):
    data=request.POST
    file=request.FILES
    file_object=request.FILES.get("avator") #获取文件名XXXX.png
    print("file对象",file_object)
    f =open(file_object,mode='wb')
    for chunk in file_object.chuncks():
        f.write(chunk)
    f.close()
    return render(request,'upload_info.html')

def depart_infomult(request):
    """批量上传 EXCEL文件"""
    from django.core.files.uploadedfile import InMemoryUploadedFile
    from openpyxl import load_workbook
    # 直接打开Excel并获取数据
     # 1,获取文件对象
    file_object=request.FILES.get("filemult")
     # 2，对象传给openpyxl，由openpyxl读取文件内容
    wb=load_workbook(file_object)
    sheet=wb.worksheets[0]
    # cell=sheet.cell(1,1) #取第一行第一列数据
    for row in sheet.iter_rows(min_row=2):
        text=row[0].value
        exists=models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart_info/")


class UpModelForm(BootstrapModelForm):
    bootstrap_exclude_fields=['imag']
    name=forms.CharField(label="姓名")
    age=forms.IntegerField(label="年龄")
    img=forms.FileField(label="头像")

def upload_form(request):
    if request.method=='GET':
        form =UpModelForm()
        return render(request,'upload_form.html',{"form":form})
    form=UpModelForm(data=request.POST,files=request.FILES)
    if form.is_valid():
        #读取到内容，自己处理每个字段的数据，返回字典，其中img-file返回的是file对象
        print(form.cleaned_data)
        #1，读取图片内容，写入到文件夹并获取文件的路径
        img_object=form.cleaned_data.get("img")

        # media_path=os.path.join(settings.MEDIA_ROOT,img_object.name)#绝对路径
        media_path = os.path.join("media", img_object.name) #相对虏路径
        f=open(media_path,mode='wb')
        for chunck in img_object.chuncks():
            f.write(chunck)
        f.close()

        #2，将文件路径写入数据库

        return JsonResponse("上传成功")
